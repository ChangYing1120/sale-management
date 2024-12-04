from django.shortcuts import render
from django.views.generic import TemplateView
from product.models import Product
from project.models import Project
from order.models import Order, ProductInOrder
from django.http import JsonResponse
import json
import csv
from django.utils.timezone import now
from django.http import HttpResponse
from datetime import datetime, timedelta
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from servicecall.models import ServiceCall

# Create your views here.
class CreateOrderView(TemplateView):
    model = Order
    template_name = "create_order.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        product_list = Product.objects.all()
        project_list = Project.objects.filter(status="active")

        context['products'] = product_list
        context['projects'] = project_list

        return context
    
class WeeklyProcurementOrderView(TemplateView):
    model = Order
    template_name = "weekly_procurement_orders.html"
    
@csrf_exempt
def ajax_create_order(request):
    if request.method == 'POST':
        try:
            # Retrieve data from the request
            order_data = json.loads(request.POST.get('order_data'))
            delivery_date = request.POST.get('delivery_date')
            project_id = request.POST.get('project_id')

            # Get the project object
            project = Project.objects.get(id=project_id)
            current_date = datetime.now().strftime('%Y-%m-%d')
            # Create the Order record
            order = Order(
                project=project,
                created_date=current_date,
                need_date=delivery_date,
                is_purchased=False
            )
            order.save()

            # Create ProjectInOrder records
            for item in order_data:
                product_id = item['product_id']
                quantity = item['quantity']
                total_price = item['totalPrice']
                discount_applied = item['discountApplied']
                discount_percentage = item['discountPercentage']

                # Get the product object
                product = Product.objects.get(id=product_id)

                # Create the ProductInOrder record
                product_in_order = ProductInOrder(
                    order=order,
                    product=product,
                    quantity=quantity,
                    customer_price=total_price,
                    discount=discount_percentage if discount_applied else 0,
                    custom_need_date=delivery_date
                )
                product_in_order.save()

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
        
# View to display weekly procurement orders
def ajax_weekly_procurement_orders(request):
    week = request.GET.get('week', datetime.now().strftime('%Y-W%U'))
    current_week = week

    # Parse the selected week and calculate its start and end date
    year, week_number = week.split('-W')
    week_start = datetime.strptime(f'{year}-W{week_number}-1', "%Y-W%U-%w").date()
    week_end = week_start + timedelta(days=6)

    # Find products that need to be ordered based on custom_need_date - delivery_time
    orders = ProductInOrder.objects.filter(custom_need_date__range=[week_start, week_end])
    
    orders_data = []
    for order in orders:
        # Calculate when the product needs to be ordered
        days_until_needed = (order.custom_need_date - datetime.today().date()).days
        if days_until_needed <= order.product.delivery_time + 7:
            orders_data.append({
                'id': order.id,
                'supplier_name': order.order.project.customer.company_name,  # Assuming project has customer relationship
                'quantity': order.quantity,
                'product_name': order.product.product_name,
                'ordered': order.order.is_purchased
            })

    # Return the data as JSON
    return JsonResponse({
        'success': True,
        'orders': orders_data,
        'current_week': current_week,
    })

# AJAX endpoint to update order status
@csrf_exempt
def update_order_status(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        ordered = request.POST.get('ordered') == 'true'
        
        try:
            order = Order.objects.get(id=order_id)
            order.ordered = ordered
            order.save()
            return JsonResponse({'success': True})
        except Order.DoesNotExist:
            return JsonResponse({'success': False})

# Export orders to Excel
def export_orders(request):
    week = request.GET.get('week')
    format_type = request.GET.get('format')

    # Logic to fetch orders based on the selected week
    orders = get_orders_by_week(week)

    if format_type == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orders_{week}.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Write headers
        ws.append(['Product', 'Supplier', 'Quantity', 'Ordered'])
        
        # Write data
        for order in orders:
            ws.append([order.product.product_name, order.order.project.customer.company_name, order.quantity, 'Yes' if order.order.is_purchased else 'No'])
        
        wb.save(response)

    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="orders_{week}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Product', 'Supplier', 'Quantity', 'Ordered'])
        
        for order in orders:
            writer.writerow([order.product.product_name, order.order.project.customer.company_name, order.quantity, 'Yes' if order.order.is_purchased else 'No'])

    return response

class ReportView(TemplateView):
    model = Order
    template_name = "report.html"

# Helper function to get orders based on the week
def get_orders_by_week(week):
    year, week_number = week.split('-W')
    week_start = datetime.datetime.strptime(f'{year}-W{week_number}-1', "%Y-W%U-%w").date()
    week_end = week_start + datetime.timedelta(days=6)
    
    # Filter orders that need to be delivered within the selected week
    # return Product.objects.filter(need_date__range=[week_start, week_end], delivery_time__lte=7)
    return ProductInOrder.objects.filter(custom_need_date__range=[week_start, week_end])

def export_report(request):
    report_type = request.GET.get('report_type')

    if report_type == 'orders':
        return export_orders_report()
    elif report_type == 'maintenance_calls':
        return export_maintenance_calls_report()
    elif report_type == 'procurement_orders':
        return export_procurement_orders_report()
    else:
        return HttpResponse(status=400)

# Export Orders Report
def export_orders_report():
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders_report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Write headers for the Excel file
    ws.append(['Order ID', 'Customer Name', 'Product', 'Quantity', 'Customer Price', 'Discount', 'Total Price', 'Order Date'])

    # Fetch all orders with related products from the database
    orders = Order.objects.prefetch_related('products_in_order__product', 'project__customer').all()

    # Write data
    for order in orders:
        for product_in_order in order.products_in_order.all():
            total_price = product_in_order.customer_price * product_in_order.quantity * (1 - product_in_order.discount/100)
            ws.append([
                order.id,
                order.project.customer.company_name,
                product_in_order.product.product_name,
                product_in_order.quantity,
                product_in_order.customer_price,
                product_in_order.discount,
                total_price,
                order.created_date
            ])

    # Save the workbook to the response
    wb.save(response)
    return response

# Export Maintenance Calls Report
def export_maintenance_calls_report():
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="maintenance_calls_report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Calls"

    # Write headers
    ws.append(['Call ID', 'Project', 'Problem Description', 'Category', 'Status', 'Tech Lead', 'Opening Date', 'Closing Date'])

    # Fetch maintenance calls from the database
    maintenance_calls = ServiceCall.objects.all()

    # Write data
    for call in maintenance_calls:
        ws.append([
            call.id,
            call.project.project_name,
            call.description,  # Changed to match the model's "description" field
            call.category,
            call.status,
            call.technical_lead.first_name,  # Added first name of technical lead
            call.opening_date,
            call.closing_date if call.closing_date else 'N/A'  # Optional closing date
        ])

    # Save the workbook to the response
    wb.save(response)
    return response


# Export Procurement Orders Report
def export_procurement_orders_report():
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="procurement_orders_report.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Orders"

    # Write headers
    ws.append(['Order ID', 'Product', 'Supplier', 'Quantity', 'Ordered Date', 'Status'])

    # Fetch procurement orders from the database
    procurement_orders = ProductInOrder.objects.select_related('product', 'order__project').all()

    # Write data
    for order in procurement_orders:
        ws.append([
            order.id,
            order.product.product_name,  # Use product relationship to get product name
            order.order.project.customer.company_name,  # Assuming order has a project and project has a customer
            order.quantity,
            order.order.created_date,  # Assuming created_date in the Order model is the ordered date
            'Ordered' if order.order.is_purchased else 'Pending'  # Assuming is_purchased marks if the order is done
        ])

    # Save the workbook to the response
    wb.save(response)
    return response